"""
Module Excel_Exporter - Export des notes annexes vers Excel

Ce module fournit la classe ExcelExporter pour exporter les notes annexes
calculées vers des fichiers Excel formatés conformes au format SYSCOHADA.

Author: Claraverse
Date: 2026-04-22
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Optional
from datetime import datetime
import os


class ExcelExporter:
    """
    Exporteur de notes annexes vers Excel avec formatage SYSCOHADA.
    
    Cette classe permet d'exporter une ou plusieurs notes annexes vers Excel
    avec un formatage professionnel incluant:
    - Bordures et couleurs conformes au format SYSCOHADA
    - Formatage numérique avec séparateurs de milliers
    - En-têtes groupés avec fusion de cellules
    - Feuilles multiples pour chaque note
    - Nom de fichier horodaté
    
    Attributes:
        fichier_sortie (str): Chemin du fichier Excel de sortie
        workbook (Workbook): Objet workbook openpyxl
    """
    
    def __init__(self, fichier_sortie: str):
        """
        Initialise l'exporteur Excel.
        
        Args:
            fichier_sortie: Chemin du fichier Excel de sortie (sans timestamp)
        """
        self.fichier_sortie = fichier_sortie
        self.workbook = Workbook()
        # Supprimer la feuille par défaut
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
    
    def exporter_note(self, 
                      numero_note: str, 
                      titre_note: str,
                      df: pd.DataFrame, 
                      colonnes_config: Dict) -> None:
        """
        Exporte une note annexe vers une feuille Excel.
        
        Args:
            numero_note: Numéro de la note (ex: "3A", "3B")
            titre_note: Titre de la note (ex: "Immobilisations Incorporelles")
            df: DataFrame contenant les données de la note
            colonnes_config: Configuration des colonnes avec structure:
                {
                    'groupes': [
                        {
                            'titre': 'VALEURS BRUTES',
                            'colonnes': ['brut_ouverture', 'augmentations', ...]
                        },
                        ...
                    ],
                    'labels': {
                        'brut_ouverture': 'Ouverture',
                        ...
                    }
                }
        """
        # Créer une nouvelle feuille
        nom_feuille = f"Note {numero_note}"
        ws = self.workbook.create_sheet(title=nom_feuille)
        
        # Ligne 1: Titre de la note
        ws.merge_cells('A1:K1')
        cell_titre = ws['A1']
        cell_titre.value = f"NOTE {numero_note} - {titre_note}"
        cell_titre.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        cell_titre.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        cell_titre.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Ligne 2: Vide pour espacement
        ws.row_dimensions[2].height = 10
        
        # Ligne 3: En-têtes de groupes
        row_idx = 3
        col_idx = 1
        
        # Colonne libellé
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = "LIBELLÉ"
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        col_idx += 1
        
        # Groupes de colonnes
        groupes = colonnes_config.get('groupes', [])
        for groupe in groupes:
            nb_colonnes = len(groupe.get('colonnes', []))
            start_col = col_idx
            end_col = col_idx + nb_colonnes - 1
            
            # Fusionner les cellules pour le groupe
            ws.merge_cells(start_row=row_idx, start_column=start_col, 
                          end_row=row_idx, end_column=end_col)
            
            cell = ws.cell(row=row_idx, column=start_col)
            cell.value = groupe['titre']
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            col_idx = end_col + 1
        
        # Ligne 4: Sous-en-têtes
        row_idx = 4
        col_idx = 1
        
        # Colonne libellé (fusionner avec ligne 3)
        ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
        
        col_idx = 2
        labels = colonnes_config.get('labels', {})
        
        for groupe in groupes:
            for colonne in groupe.get('colonnes', []):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = labels.get(colonne, colonne)
                cell.font = Font(name='Arial', size=10, bold=False, color='FFFFFF')
                cell.fill = PatternFill(start_color='7F8C8D', end_color='7F8C8D', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                col_idx += 1
        
        # Lignes de données
        row_idx = 5
        colonnes_ordre = []
        for groupe in groupes:
            colonnes_ordre.extend(groupe.get('colonnes', []))
        
        for _, row_data in df.iterrows():
            col_idx = 1
            
            # Colonne libellé
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data.get('libelle', '')
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Déterminer si c'est une ligne de total
            is_total = 'TOTAL' in str(row_data.get('libelle', '')).upper()
            
            if is_total:
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
            
            col_idx += 1
            
            # Colonnes de montants
            for colonne in colonnes_ordre:
                cell = ws.cell(row=row_idx, column=col_idx)
                montant = row_data.get(colonne, 0.0)
                
                if pd.notna(montant) and montant != 0:
                    cell.value = montant
                    cell.number_format = '#,##0'  # Format avec séparateur de milliers
                else:
                    cell.value = '-'
                
                cell.font = Font(name='Arial', size=10, bold=is_total)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                
                if is_total:
                    cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
                
                col_idx += 1
            
            row_idx += 1
        
        # Appliquer le formatage (bordures, largeurs de colonnes)
        self.appliquer_formatage(ws, row_idx - 1, len(colonnes_ordre) + 1)
    
    def exporter_toutes_notes(self, notes_data: Dict[str, Dict]) -> None:
        """
        Exporte toutes les notes annexes vers le fichier Excel.
        
        Args:
            notes_data: Dictionnaire avec structure:
                {
                    '3A': {
                        'titre': 'Immobilisations Incorporelles',
                        'df': DataFrame,
                        'colonnes_config': Dict
                    },
                    '3B': {...},
                    ...
                }
        """
        for numero_note, note_info in notes_data.items():
            self.exporter_note(
                numero_note=numero_note,
                titre_note=note_info['titre'],
                df=note_info['df'],
                colonnes_config=note_info['colonnes_config']
            )
    
    def appliquer_formatage(self, ws, max_row: int, max_col: int) -> None:
        """
        Applique le formatage aux cellules (bordures, largeurs).
        
        Args:
            ws: Feuille de calcul openpyxl
            max_row: Nombre maximum de lignes
            max_col: Nombre maximum de colonnes
        """
        # Définir les bordures
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Appliquer les bordures à toutes les cellules de données
        for row in range(3, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
        
        # Ajuster les largeurs de colonnes
        from openpyxl.utils import get_column_letter
        
        ws.column_dimensions['A'].width = 40  # Colonne libellé plus large
        
        for col in range(2, max_col + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
    
    def sauvegarder(self, avec_timestamp: bool = True) -> str:
        """
        Sauvegarde le fichier Excel avec nom horodaté.
        
        Args:
            avec_timestamp: Si True, ajoute un timestamp au nom du fichier
        
        Returns:
            Chemin complet du fichier sauvegardé
        """
        if avec_timestamp:
            # Extraire le nom de base et l'extension
            base_name, ext = os.path.splitext(self.fichier_sortie)
            
            # Ajouter le timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fichier_final = f"{base_name}_{timestamp}{ext}"
        else:
            fichier_final = self.fichier_sortie
        
        # Créer le répertoire si nécessaire
        os.makedirs(os.path.dirname(fichier_final) if os.path.dirname(fichier_final) else '.', exist_ok=True)
        
        # Sauvegarder le workbook
        self.workbook.save(fichier_final)
        
        return fichier_final


# Exemple d'utilisation
if __name__ == "__main__":
    # Créer des données de test
    data_note_3a = {
        'libelle': [
            'Frais de recherche et de développement',
            'Brevets, licences, logiciels',
            'Fonds commercial et droit au bail',
            'Autres immobilisations incorporelles',
            'TOTAL IMMOBILISATIONS INCORPORELLES'
        ],
        'brut_ouverture': [1500000, 2500000, 5000000, 800000, 9800000],
        'augmentations': [500000, 300000, 0, 200000, 1000000],
        'diminutions': [200000, 0, 0, 100000, 300000],
        'brut_cloture': [1800000, 2800000, 5000000, 900000, 10500000],
        'amort_ouverture': [300000, 1000000, 0, 400000, 1700000],
        'dotations': [200000, 300000, 0, 100000, 600000],
        'reprises': [50000, 0, 0, 50000, 100000],
        'amort_cloture': [450000, 1300000, 0, 450000, 2200000],
        'vnc_ouverture': [1200000, 1500000, 5000000, 400000, 8100000],
        'vnc_cloture': [1350000, 1500000, 5000000, 450000, 8300000]
    }
    
    df_3a = pd.DataFrame(data_note_3a)
    
    # Configuration des colonnes
    colonnes_config = {
        'groupes': [
            {
                'titre': 'VALEURS BRUTES',
                'colonnes': ['brut_ouverture', 'augmentations', 'diminutions', 'brut_cloture']
            },
            {
                'titre': 'AMORTISSEMENTS',
                'colonnes': ['amort_ouverture', 'dotations', 'reprises', 'amort_cloture']
            },
            {
                'titre': 'VALEURS NETTES COMPTABLES',
                'colonnes': ['vnc_ouverture', 'vnc_cloture']
            }
        ],
        'labels': {
            'brut_ouverture': 'Ouverture',
            'augmentations': 'Augmentations',
            'diminutions': 'Diminutions',
            'brut_cloture': 'Clôture',
            'amort_ouverture': 'Ouverture',
            'dotations': 'Dotations',
            'reprises': 'Reprises',
            'amort_cloture': 'Clôture',
            'vnc_ouverture': 'Ouverture',
            'vnc_cloture': 'Clôture'
        }
    }
    
    # Créer l'exporteur
    exporter = ExcelExporter('../Tests/notes_annexes_test.xlsx')
    
    # Exporter la note 3A
    exporter.exporter_note(
        numero_note='3A',
        titre_note='Immobilisations Incorporelles',
        df=df_3a,
        colonnes_config=colonnes_config
    )
    
    # Sauvegarder le fichier
    fichier_sauvegarde = exporter.sauvegarder(avec_timestamp=True)
    
    print(f"✓ Fichier Excel généré avec succès: {fichier_sauvegarde}")
    print(f"✓ Nombre de feuilles: {len(exporter.workbook.sheetnames)}")
    print(f"✓ Feuilles: {exporter.workbook.sheetnames}")
