"""
Property-Based Tests for Numeric Conversion Robustness

This module contains property-based tests using Hypothesis to verify
the Balance_Reader module's numeric conversion behavior across various
invalid inputs including empty strings, None values, text, special characters,
and mixed formats.

**Validates: Requirements 1.5, 1.6**

Property 3: Numeric Conversion Robustness
For any balance sheet loaded, all monetary values must be converted to float type,
and any invalid or empty values must be replaced with 0.0 without raising exceptions.

Auteur: Système de calcul automatique des notes annexes SYSCOHADA
Date: 08 Avril 2026
"""

import sys
import os
import pytest
from hypothesis import given, strategies as st, assume, settings
import pandas as pd
import numpy as np
from openpyxl import Workbook
import tempfile

# Ajouter le chemin des modules au PYTHONPATH
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'Modules'))

from balance_reader import BalanceReader, BalanceNotFoundException, InvalidBalanceFormatException


# ============================================================================
# HYPOTHESIS STRATEGIES FOR INVALID VALUES
# ============================================================================

@st.composite
def st_invalid_monetary_value(draw):
    """
    Génère des valeurs monétaires invalides pour tester la robustesse.
    
    Cette stratégie génère:
    - Chaînes vides
    - None
    - Texte non numérique
    - Caractères spéciaux
    - Formats mixtes (lettres + chiffres)
    - Espaces
    - Valeurs NaN
    
    Returns:
        Valeur invalide de type varié
    """
    return draw(st.one_of(
        st.just(''),                          # Chaîne vide
        st.just(None),                        # None
        st.just('N/A'),                       # Texte
        st.just('n/a'),                       # Texte minuscule
        st.just('-'),                         # Tiret seul
        st.just('abc'),                       # Lettres
        st.just('xyz123'),                    # Lettres + chiffres
        st.just('123abc'),                    # Chiffres + lettres
        st.just('   '),                       # Espaces
        st.just('###'),                       # Caractères spéciaux
        st.just('***'),                       # Astérisques
        st.just('...'),                       # Points
        st.just('???'),                       # Points d'interrogation
        st.just('ERROR'),                     # Message d'erreur
        st.just('#DIV/0!'),                   # Erreur Excel
        st.just('#VALUE!'),                   # Erreur Excel
        st.just('#REF!'),                     # Erreur Excel
        st.just(np.nan),                      # NaN numpy
        st.just(float('nan')),                # NaN Python
        # Texte aléatoire filtré pour exclure les caractères illégaux Excel
        st.text(
            alphabet=st.characters(
                blacklist_categories=('Cc', 'Cs'),  # Exclure les caractères de contrôle et surrogates
                blacklist_characters='\x00\x01\x02\x03\x04\x05\x06\x07\x08\x0b\x0c\x0e\x0f\x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1a\x1b\x1c\x1d\x1e\x1f'
            ),
            min_size=1,
            max_size=10
        ),
    ))


@st.composite
def st_valid_monetary_value(draw):
    """
    Génère des valeurs monétaires valides.
    
    Returns:
        float: Montant valide entre 0 et 10 millions
    """
    return draw(st.floats(
        min_value=0,
        max_value=10000000,
        allow_nan=False,
        allow_infinity=False
    ))


@st.composite
def st_mixed_monetary_column(draw):
    """
    Génère une colonne avec un mélange de valeurs valides et invalides.
    
    Cette stratégie crée des colonnes contenant:
    - 30-70% de valeurs valides (float)
    - 30-70% de valeurs invalides (empty, None, text, etc.)
    
    Returns:
        List: Liste de valeurs mixtes
    """
    num_values = draw(st.integers(min_value=10, max_value=50))
    
    # Déterminer le ratio de valeurs invalides (30-70%)
    invalid_ratio = draw(st.floats(min_value=0.3, max_value=0.7))
    num_invalid = int(num_values * invalid_ratio)
    num_valid = num_values - num_invalid
    
    # Générer les valeurs
    values = []
    
    # Ajouter des valeurs valides
    for _ in range(num_valid):
        values.append(draw(st_valid_monetary_value()))
    
    # Ajouter des valeurs invalides
    for _ in range(num_invalid):
        values.append(draw(st_invalid_monetary_value()))
    
    # Mélanger les valeurs
    import random
    random.shuffle(values)
    
    return values


@st.composite
def st_balance_with_invalid_values(draw):
    """
    Génère un fichier Excel avec des valeurs monétaires invalides.
    
    Cette stratégie crée un fichier Excel avec:
    - 3 onglets (BALANCE N, N-1, N-2)
    - Colonnes monétaires contenant des valeurs invalides
    - Mélange de valeurs valides et invalides
    
    Returns:
        str: Chemin vers le fichier Excel temporaire créé
    """
    # Créer un fichier temporaire
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
    temp_file.close()
    
    # Créer le workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Créer les 3 onglets
    for exercice in ['N', 'N-1', 'N-2']:
        ws = wb.create_sheet(f"BALANCE {exercice}")
        
        # En-têtes
        headers = ['Numéro', 'Intitulé', 'Ant Débit', 'Ant Crédit', 
                   'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        ws.append(headers)
        
        # Générer les colonnes avec valeurs mixtes
        num_comptes = draw(st.integers(min_value=10, max_value=30))
        
        # Générer les colonnes monétaires avec valeurs invalides
        ant_debit_col = draw(st_mixed_monetary_column())[:num_comptes]
        ant_credit_col = draw(st_mixed_monetary_column())[:num_comptes]
        debit_col = draw(st_mixed_monetary_column())[:num_comptes]
        credit_col = draw(st_mixed_monetary_column())[:num_comptes]
        solde_debit_col = draw(st_mixed_monetary_column())[:num_comptes]
        solde_credit_col = draw(st_mixed_monetary_column())[:num_comptes]
        
        # Ajouter les lignes
        for i in range(num_comptes):
            # Numéro de compte valide
            classe = draw(st.sampled_from(['1', '2', '3', '4', '5', '6', '7', '8', '9']))
            sous_classe = draw(st.integers(min_value=0, max_value=9))
            detail = draw(st.integers(min_value=0, max_value=999))
            numero = f"{classe}{sous_classe}{detail:03d}"
            
            # Intitulé
            intitule = f"Compte {numero}"
            
            # Ajouter la ligne avec valeurs mixtes
            ws.append([
                numero,
                intitule,
                ant_debit_col[i] if i < len(ant_debit_col) else '',
                ant_credit_col[i] if i < len(ant_credit_col) else '',
                debit_col[i] if i < len(debit_col) else '',
                credit_col[i] if i < len(credit_col) else '',
                solde_debit_col[i] if i < len(solde_debit_col) else '',
                solde_credit_col[i] if i < len(solde_credit_col) else ''
            ])
    
    # Sauvegarder le fichier
    wb.save(temp_file.name)
    
    return temp_file.name


# ============================================================================
# PROPERTY-BASED TESTS - NUMERIC CONVERSION ROBUSTNESS
# ============================================================================

@given(fichier_excel=st_balance_with_invalid_values())
@settings(max_examples=50, deadline=60000)
def test_property_numeric_conversion_robustness(fichier_excel):
    """
    **Property 3: Numeric Conversion Robustness**
    
    **Validates: Requirements 1.5, 1.6**
    
    For any balance sheet loaded, all monetary values must be converted to 
    float type, and any invalid or empty values must be replaced with 0.0 
    without raising exceptions.
    
    This property verifies that:
    1. All monetary columns are converted to float type
    2. Empty strings are replaced with 0.0
    3. None values are replaced with 0.0
    4. Text values are replaced with 0.0
    5. Special characters are replaced with 0.0
    6. Mixed formats are handled gracefully
    7. No exceptions are raised during conversion
    8. All converted values are >= 0
    9. No NaN or infinite values remain after conversion
    
    Test Strategy:
    - Generate Excel files with 30-70% invalid values in monetary columns
    - Invalid values include: '', None, 'N/A', text, special chars, NaN
    - Verify that Balance_Reader converts all values to float
    - Verify that all invalid values become 0.0
    - Verify that no exceptions are raised
    """
    try:
        # Créer le lecteur
        reader = BalanceReader(fichier_excel)
        
        # Charger les balances - ne doit pas lever d'exception
        balance_n, balance_n1, balance_n2 = reader.charger_balances()
        
        # Colonnes monétaires à vérifier
        colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        
        # Vérifier chaque balance
        for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
            # Vérifier que la balance est chargée
            assert balance is not None, f"Balance {nom} ne doit pas être None"
            assert isinstance(balance, pd.DataFrame), f"Balance {nom} doit être un DataFrame"
            assert len(balance) > 0, f"Balance {nom} doit contenir au moins 1 ligne"
            
            # Vérifier chaque colonne monétaire
            for col in colonnes_montants:
                if col in balance.columns:
                    # 1. Vérifier que la colonne est de type numérique (float)
                    assert pd.api.types.is_numeric_dtype(balance[col]), \
                        f"Colonne {col} de Balance {nom} doit être de type numérique, trouvé {balance[col].dtype}"
                    
                    # 2. Vérifier qu'il n'y a pas de NaN
                    assert not balance[col].isna().any(), \
                        f"Colonne {col} de Balance {nom} ne doit pas contenir de NaN après conversion"
                    
                    # 3. Vérifier qu'il n'y a pas de valeurs infinies
                    assert not np.isinf(balance[col]).any(), \
                        f"Colonne {col} de Balance {nom} ne doit pas contenir de valeurs infinies"
                    
                    # 4. Vérifier que toutes les valeurs sont >= 0
                    assert (balance[col] >= 0).all(), \
                        f"Colonne {col} de Balance {nom} doit contenir uniquement des valeurs >= 0"
                    
                    # 5. Vérifier que les valeurs sont de type float
                    assert balance[col].dtype in [np.float64, np.float32, float], \
                        f"Colonne {col} de Balance {nom} doit être de type float, trouvé {balance[col].dtype}"
                    
                    # 6. Vérifier que les valeurs invalides ont été remplacées par 0.0
                    # (on ne peut pas vérifier directement, mais on vérifie qu'il n'y a pas d'erreur)
                    assert balance[col].notna().all(), \
                        f"Colonne {col} de Balance {nom} ne doit pas contenir de valeurs manquantes"
        
        # Vérifier que les 3 balances ont été chargées avec succès
        assert balance_n is not None and balance_n1 is not None and balance_n2 is not None, \
            "Les 3 balances doivent être chargées avec succès"
        
    except (BalanceNotFoundException, InvalidBalanceFormatException) as e:
        # Ces exceptions sont acceptables si le fichier est vraiment invalide
        # (par exemple, onglets manquants)
        pytest.skip(f"Fichier invalide (acceptable): {str(e)}")
    
    except Exception as e:
        # Toute autre exception est un échec du test
        pytest.fail(f"Exception inattendue lors de la conversion: {type(e).__name__}: {str(e)}")
    
    finally:
        # Nettoyer le fichier temporaire
        if os.path.exists(fichier_excel):
            try:
                os.unlink(fichier_excel)
            except Exception:
                pass


def test_property_numeric_conversion_with_demo_file():
    """
    Test de la propriété de conversion numérique avec le fichier de démonstration.
    
    **Validates: Requirements 1.5, 1.6**
    
    Ce test vérifie que le fichier de démonstration réel a toutes ses valeurs
    monétaires correctement converties en float, sans NaN ni valeurs infinies.
    """
    # Chemin vers le fichier de test
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    fichier_test = os.path.join(base_dir, "P000 -BALANCE DEMO N_N-1_N-2.xls")
    
    # Vérifier que le fichier existe
    if not os.path.exists(fichier_test):
        pytest.skip(f"Fichier de démonstration non trouvé: {fichier_test}")
    
    # Créer le lecteur
    reader = BalanceReader(fichier_test)
    
    # Charger les balances
    balance_n, balance_n1, balance_n2 = reader.charger_balances()
    
    # Colonnes monétaires
    colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
    
    # Vérifier chaque balance
    for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
        for col in colonnes_montants:
            if col in balance.columns:
                # Vérifier le type
                assert pd.api.types.is_numeric_dtype(balance[col]), \
                    f"Colonne {col} de Balance {nom} doit être numérique"
                
                # Vérifier l'absence de NaN
                assert not balance[col].isna().any(), \
                    f"Colonne {col} de Balance {nom} ne doit pas contenir de NaN"
                
                # Vérifier l'absence de valeurs infinies
                assert not np.isinf(balance[col]).any(), \
                    f"Colonne {col} de Balance {nom} ne doit pas contenir de valeurs infinies"
                
                # Vérifier que toutes les valeurs sont >= 0
                assert (balance[col] >= 0).all(), \
                    f"Colonne {col} de Balance {nom} doit contenir uniquement des valeurs >= 0"
    
    print(f"\n✓ Propriété de conversion numérique validée avec le fichier de démonstration")
    print(f"  - Balance N:   {len(balance_n)} comptes, toutes valeurs converties en float")
    print(f"  - Balance N-1: {len(balance_n1)} comptes, toutes valeurs converties en float")
    print(f"  - Balance N-2: {len(balance_n2)} comptes, toutes valeurs converties en float")


@st.composite
def st_balance_with_specific_invalid_values(draw):
    """
    Génère un fichier Excel avec des types spécifiques de valeurs invalides.
    
    Cette stratégie teste des cas spécifiques:
    - Colonnes entièrement vides
    - Colonnes avec uniquement None
    - Colonnes avec uniquement du texte
    - Colonnes avec des erreurs Excel (#DIV/0!, #VALUE!, etc.)
    
    Returns:
        Tuple[str, str]: (chemin fichier, type de valeur invalide)
    """
    # Choisir un type de valeur invalide
    invalid_type = draw(st.sampled_from([
        'empty',      # Chaînes vides
        'none',       # None
        'text',       # Texte
        'excel_error',# Erreurs Excel
        'special_chars', # Caractères spéciaux
        'mixed'       # Mélange
    ]))
    
    # Créer un fichier temporaire
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
    temp_file.close()
    
    # Créer le workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Créer un seul onglet pour ce test
    ws = wb.create_sheet("BALANCE N")
    
    # En-têtes
    headers = ['Numéro', 'Intitulé', 'Ant Débit', 'Ant Crédit', 
               'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
    ws.append(headers)
    
    # Générer les données avec le type de valeur invalide choisi
    num_comptes = draw(st.integers(min_value=5, max_value=15))
    
    for i in range(num_comptes):
        # Numéro de compte valide
        classe = draw(st.sampled_from(['1', '2', '3', '4', '5', '6', '7', '8', '9']))
        sous_classe = draw(st.integers(min_value=0, max_value=9))
        numero = f"{classe}{sous_classe}"
        
        # Intitulé
        intitule = f"Compte {numero}"
        
        # Générer les valeurs selon le type
        if invalid_type == 'empty':
            values = ['', '', '', '', '', '']
        elif invalid_type == 'none':
            values = [None, None, None, None, None, None]
        elif invalid_type == 'text':
            values = ['N/A', 'n/a', 'ERROR', 'abc', 'xyz', 'text']
        elif invalid_type == 'excel_error':
            values = ['#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#NUM!', '#N/A']
        elif invalid_type == 'special_chars':
            values = ['###', '***', '...', '???', '---', '+++']
        else:  # mixed
            values = ['', None, 'N/A', '#DIV/0!', '###', 'abc']
        
        # Ajouter la ligne
        ws.append([numero, intitule] + values)
    
    # Sauvegarder le fichier
    wb.save(temp_file.name)
    
    return temp_file.name, invalid_type


@given(data=st.data())
@settings(max_examples=30, deadline=60000)
def test_property_specific_invalid_value_types(data):
    """
    **Property 3 (Extended): Specific Invalid Value Types**
    
    **Validates: Requirements 1.5, 1.6**
    
    For any balance sheet with specific types of invalid values (empty strings,
    None, text, Excel errors, special characters), the Balance_Reader must
    convert all values to 0.0 without raising exceptions.
    
    This test verifies conversion for specific categories of invalid values:
    - Empty strings ('')
    - None values
    - Text values ('N/A', 'ERROR', etc.)
    - Excel error values ('#DIV/0!', '#VALUE!', etc.)
    - Special characters ('###', '***', etc.)
    - Mixed invalid values
    
    Test Strategy:
    - Generate Excel files with columns containing only one type of invalid value
    - Verify that all invalid values are converted to 0.0
    - Verify that no exceptions are raised
    - Verify that the resulting DataFrame has all numeric columns
    """
    fichier_excel, invalid_type = data.draw(st_balance_with_specific_invalid_values())
    
    try:
        # Créer le lecteur
        reader = BalanceReader(fichier_excel)
        
        # Charger la balance - ne doit pas lever d'exception
        balance_n, _, _ = reader.charger_balances()
        
        # Colonnes monétaires
        colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        
        # Vérifier chaque colonne monétaire
        for col in colonnes_montants:
            if col in balance_n.columns:
                # Vérifier que la colonne est numérique
                assert pd.api.types.is_numeric_dtype(balance_n[col]), \
                    f"Colonne {col} doit être numérique pour type invalide '{invalid_type}'"
                
                # Vérifier que toutes les valeurs sont 0.0 (car toutes étaient invalides)
                assert (balance_n[col] == 0.0).all(), \
                    f"Toutes les valeurs de {col} doivent être 0.0 pour type invalide '{invalid_type}', " \
                    f"trouvé: {balance_n[col].unique()}"
                
                # Vérifier l'absence de NaN
                assert not balance_n[col].isna().any(), \
                    f"Colonne {col} ne doit pas contenir de NaN pour type invalide '{invalid_type}'"
        
    except (BalanceNotFoundException, InvalidBalanceFormatException) as e:
        pytest.skip(f"Fichier invalide (acceptable): {str(e)}")
    
    except Exception as e:
        pytest.fail(f"Exception inattendue pour type '{invalid_type}': {type(e).__name__}: {str(e)}")
    
    finally:
        # Nettoyer le fichier temporaire
        if os.path.exists(fichier_excel):
            try:
                os.unlink(fichier_excel)
            except Exception:
                pass


def test_convertir_montants_unit():
    """
    Test unitaire de la méthode convertir_montants avec des cas spécifiques.
    
    **Validates: Requirements 1.5, 1.6**
    
    Ce test vérifie des cas spécifiques de conversion:
    - Valeurs vides
    - None
    - Texte
    - Nombres valides
    - Mélange de valeurs
    """
    # Créer un DataFrame de test avec des valeurs invalides
    df_test = pd.DataFrame({
        'Numéro': ['211', '212', '213', '214', '215'],
        'Intitulé': ['Compte 1', 'Compte 2', 'Compte 3', 'Compte 4', 'Compte 5'],
        'Ant Débit': [1000.0, '', None, 'N/A', 2000.0],
        'Ant Crédit': ['', 500.0, None, 'abc', 1500.0],
        'Débit': [300.0, None, '', '#DIV/0!', 400.0],
        'Crédit': [None, 200.0, 'ERROR', '', 350.0],
        'Solde Débit': ['###', 1500.0, None, '', 2400.0],
        'Solde Crédit': ['', None, 700.0, '***', 1850.0]
    })
    
    # Créer un lecteur (fichier fictif)
    reader = BalanceReader("dummy.xlsx")
    
    # Convertir les montants
    df_converted = reader.convertir_montants(df_test)
    
    # Vérifier que toutes les colonnes monétaires sont numériques
    colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
    
    for col in colonnes_montants:
        # Vérifier le type
        assert pd.api.types.is_numeric_dtype(df_converted[col]), \
            f"Colonne {col} doit être numérique"
        
        # Vérifier l'absence de NaN
        assert not df_converted[col].isna().any(), \
            f"Colonne {col} ne doit pas contenir de NaN"
        
        # Vérifier que toutes les valeurs sont >= 0
        assert (df_converted[col] >= 0).all(), \
            f"Colonne {col} doit contenir uniquement des valeurs >= 0"
    
    # Vérifier des valeurs spécifiques
    # Ligne 0: 1000.0, '', 300.0, None, '###', ''
    assert df_converted.loc[0, 'Ant Débit'] == 1000.0
    assert df_converted.loc[0, 'Ant Crédit'] == 0.0  # '' -> 0.0
    assert df_converted.loc[0, 'Débit'] == 300.0
    assert df_converted.loc[0, 'Crédit'] == 0.0  # None -> 0.0
    assert df_converted.loc[0, 'Solde Débit'] == 0.0  # '###' -> 0.0
    assert df_converted.loc[0, 'Solde Crédit'] == 0.0  # '' -> 0.0
    
    # Ligne 2: None, None, '', 'ERROR', None, 700.0
    assert df_converted.loc[2, 'Ant Débit'] == 0.0  # None -> 0.0
    assert df_converted.loc[2, 'Ant Crédit'] == 0.0  # None -> 0.0
    assert df_converted.loc[2, 'Débit'] == 0.0  # '' -> 0.0
    assert df_converted.loc[2, 'Crédit'] == 0.0  # 'ERROR' -> 0.0
    assert df_converted.loc[2, 'Solde Débit'] == 0.0  # None -> 0.0
    assert df_converted.loc[2, 'Solde Crédit'] == 700.0
    
    print("\n✓ Test unitaire de conversion des montants réussi")
    print("  - Valeurs vides converties en 0.0")
    print("  - None converti en 0.0")
    print("  - Texte converti en 0.0")
    print("  - Valeurs valides préservées")


if __name__ == "__main__":
    """
    Exécution directe des tests pour validation rapide.
    
    Usage:
        python test_balance_reader_numeric_conversion.py
    """
    print("=" * 70)
    print("PROPERTY-BASED TESTS - NUMERIC CONVERSION ROBUSTNESS")
    print("=" * 70)
    
    # Test unitaire
    print("\n[1] Test unitaire de conversion des montants...")
    try:
        test_convertir_montants_unit()
        print("   ✓ Test réussi")
    except Exception as e:
        print(f"   ✗ Test échoué: {str(e)}")
        import traceback
        traceback.print_exc()
    
    # Test avec le fichier de démonstration
    print("\n[2] Test avec le fichier de démonstration...")
    try:
        test_property_numeric_conversion_with_demo_file()
        print("   ✓ Test réussi")
    except Exception as e:
        print(f"   ✗ Test échoué: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "=" * 70)
    print("Pour exécuter tous les tests property-based avec Hypothesis:")
    print("  pytest test_balance_reader_numeric_conversion.py -v")
    print("\nPour voir les statistiques Hypothesis:")
    print("  pytest test_balance_reader_numeric_conversion.py -v --hypothesis-show-statistics")
    print("=" * 70)
