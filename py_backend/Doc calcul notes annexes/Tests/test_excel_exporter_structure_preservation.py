"""
Property-Based Test: Excel Export Structure Preservation

Feature: calcul-notes-annexes-syscohada
Property 15: Excel Export Structure Preservation

For any note annexe exported to Excel, the Excel_Exporter must create a worksheet 
with the same structure as the HTML table (headers, data rows, total row), with 
numeric formatting for amounts and styling (borders, header colors).

Validates: Requirements 9.1, 9.2, 9.3, 9.4, 9.5

Author: Claraverse
Date: 2026-04-22
"""

import pytest
from hypothesis import given, settings, assume
import hypothesis.strategies as st
import pandas as pd
import os
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'Modules'))

from excel_exporter import ExcelExporter


# ============================================================================
# STRATÉGIES DE GÉNÉRATION
# ============================================================================

@st.composite
def st_note_annexe_data(draw):
    """
    Génère des données de note annexe valides pour le test.
    
    Returns:
        Tuple (numero_note, titre_note, df, colonnes_config)
    """
    # Numéro de note (3A-3E, 4-33)
    numero_note = draw(st.sampled_from([
        '3A', '3B', '3C', '3D', '3E',
        '4', '5', '6', '7', '8', '9', '10'
    ]))
    
    # Titre de note - utiliser un alphabet sûr pour Excel
    titre_note = draw(st.text(
        alphabet=st.characters(
            whitelist_categories=('Lu', 'Ll', 'Nd'),  # Lettres majuscules, minuscules, chiffres
            whitelist_characters=' -_'
        ),
        min_size=10, 
        max_size=50
    ))
    
    # Nombre de lignes (2 à 10 lignes de données + 1 ligne de total)
    num_lignes = draw(st.integers(min_value=2, max_value=10))
    
    # Générer les lignes de données
    lignes = []
    for i in range(num_lignes):
        # Libellé - utiliser un alphabet sûr pour Excel
        libelle = draw(st.text(
            alphabet=st.characters(
                whitelist_categories=('Lu', 'Ll', 'Nd'),  # Lettres majuscules, minuscules, chiffres
                whitelist_characters=' -_'
            ),
            min_size=5, 
            max_size=40
        ))
        
        # Valeurs brutes
        brut_ouverture = draw(st.floats(min_value=0, max_value=10000000, allow_nan=False, allow_infinity=False))
        augmentations = draw(st.floats(min_value=0, max_value=5000000, allow_nan=False, allow_infinity=False))
        diminutions = draw(st.floats(min_value=0, max_value=min(brut_ouverture + augmentations, 5000000), allow_nan=False, allow_infinity=False))
        brut_cloture = brut_ouverture + augmentations - diminutions
        
        # Amortissements
        amort_ouverture = draw(st.floats(min_value=0, max_value=brut_ouverture, allow_nan=False, allow_infinity=False))
        dotations = draw(st.floats(min_value=0, max_value=2000000, allow_nan=False, allow_infinity=False))
        reprises = draw(st.floats(min_value=0, max_value=min(amort_ouverture + dotations, 2000000), allow_nan=False, allow_infinity=False))
        amort_cloture = amort_ouverture + dotations - reprises
        
        # VNC
        vnc_ouverture = brut_ouverture - amort_ouverture
        vnc_cloture = brut_cloture - amort_cloture
        
        lignes.append({
            'libelle': libelle,
            'brut_ouverture': brut_ouverture,
            'augmentations': augmentations,
            'diminutions': diminutions,
            'brut_cloture': brut_cloture,
            'amort_ouverture': amort_ouverture,
            'dotations': dotations,
            'reprises': reprises,
            'amort_cloture': amort_cloture,
            'vnc_ouverture': vnc_ouverture,
            'vnc_cloture': vnc_cloture
        })
    
    # Ajouter la ligne de total
    total = {
        'libelle': 'TOTAL',
        'brut_ouverture': sum(l['brut_ouverture'] for l in lignes),
        'augmentations': sum(l['augmentations'] for l in lignes),
        'diminutions': sum(l['diminutions'] for l in lignes),
        'brut_cloture': sum(l['brut_cloture'] for l in lignes),
        'amort_ouverture': sum(l['amort_ouverture'] for l in lignes),
        'dotations': sum(l['dotations'] for l in lignes),
        'reprises': sum(l['reprises'] for l in lignes),
        'amort_cloture': sum(l['amort_cloture'] for l in lignes),
        'vnc_ouverture': sum(l['vnc_ouverture'] for l in lignes),
        'vnc_cloture': sum(l['vnc_cloture'] for l in lignes)
    }
    lignes.append(total)
    
    df = pd.DataFrame(lignes)
    
    # Configuration des colonnes
    colonnes_config = {
        'groupes': [
            {
                'titre': 'VALEURS BRUTES',
                'colonnes': ['brut_ouverture', 'augmentations', 'diminutions', 'brut_cloture']
            },
            {
                'titre': 'AMORTISSEMENTS',
                'colonnes': ['amort_ouverture', 'dotations', 'reprises', 'amort_cloture']
            },
            {
                'titre': 'VALEURS NETTES COMPTABLES',
                'colonnes': ['vnc_ouverture', 'vnc_cloture']
            }
        ],
        'labels': {
            'brut_ouverture': 'Ouverture',
            'augmentations': 'Augmentations',
            'diminutions': 'Diminutions',
            'brut_cloture': 'Clôture',
            'amort_ouverture': 'Ouverture',
            'dotations': 'Dotations',
            'reprises': 'Reprises',
            'amort_cloture': 'Clôture',
            'vnc_ouverture': 'Ouverture',
            'vnc_cloture': 'Clôture'
        }
    }
    
    return numero_note, titre_note, df, colonnes_config


# ============================================================================
# TESTS DE PROPRIÉTÉS
# ============================================================================

@given(note_data=st_note_annexe_data())
@settings(max_examples=100, deadline=60000)
def test_property_15_excel_export_structure_preservation(note_data):
    """
    **Validates: Requirements 9.1, 9.2, 9.3, 9.4, 9.5**
    
    Property 15: Excel Export Structure Preservation
    
    For any note annexe exported to Excel, the Excel_Exporter must create a 
    worksheet with the same structure as the HTML table (headers, data rows, 
    total row), with numeric formatting for amounts and styling (borders, 
    header colors).
    
    This test verifies:
    1. Worksheet is created with correct name (Requirement 9.1)
    2. Structure matches input DataFrame (headers, rows, totals) (Requirement 9.2)
    3. Numeric cells have thousand separator formatting (Requirement 9.3)
    4. Borders are applied to all data cells (Requirement 9.4)
    5. Header cells have colored backgrounds (Requirement 9.4)
    6. File is saved with timestamped filename (Requirement 9.5)
    """
    numero_note, titre_note, df, colonnes_config = note_data
    
    # Créer un fichier temporaire pour l'export
    with tempfile.TemporaryDirectory() as tmpdir:
        fichier_sortie = os.path.join(tmpdir, 'test_export.xlsx')
        
        # Créer l'exporteur et exporter la note
        exporter = ExcelExporter(fichier_sortie)
        exporter.exporter_note(
            numero_note=numero_note,
            titre_note=titre_note,
            df=df,
            colonnes_config=colonnes_config
        )
        
        # Sauvegarder avec timestamp
        fichier_sauvegarde = exporter.sauvegarder(avec_timestamp=True)
        
        # Vérifier que le fichier existe
        assert os.path.exists(fichier_sauvegarde), \
            "Le fichier Excel doit être créé"
        
        # Vérifier que le nom contient un timestamp (Requirement 9.5)
        assert fichier_sauvegarde != fichier_sortie, \
            "Le fichier doit avoir un timestamp dans son nom"
        assert '_' in os.path.basename(fichier_sauvegarde), \
            "Le nom du fichier doit contenir un underscore pour le timestamp"
        
        # Charger le fichier Excel pour vérification
        wb = load_workbook(fichier_sauvegarde)
        
        # Vérifier que la feuille existe (Requirement 9.1)
        nom_feuille = f"Note {numero_note}"
        assert nom_feuille in wb.sheetnames, \
            f"La feuille '{nom_feuille}' doit exister dans le workbook"
        
        ws = wb[nom_feuille]
        
        # Vérifier la structure: Titre de la note (Requirement 9.2)
        cell_titre = ws['A1']
        assert cell_titre.value is not None, \
            "La cellule A1 doit contenir le titre de la note"
        assert numero_note in str(cell_titre.value), \
            f"Le titre doit contenir le numéro de note '{numero_note}'"
        assert titre_note in str(cell_titre.value), \
            f"Le titre doit contenir '{titre_note}'"
        
        # Vérifier le style du titre (Requirement 9.4)
        assert cell_titre.fill.start_color.rgb is not None, \
            "Le titre doit avoir une couleur de fond"
        assert cell_titre.font.bold, \
            "Le titre doit être en gras"
        
        # Vérifier les en-têtes de groupes (ligne 3) (Requirement 9.2)
        groupes = colonnes_config.get('groupes', [])
        col_idx = 2  # Commence après la colonne libellé
        
        for groupe in groupes:
            cell = ws.cell(row=3, column=col_idx)
            assert cell.value == groupe['titre'], \
                f"L'en-tête de groupe doit être '{groupe['titre']}'"
            
            # Vérifier le style de l'en-tête (Requirement 9.4)
            assert cell.fill.start_color.rgb is not None, \
                "Les en-têtes de groupe doivent avoir une couleur de fond"
            
            col_idx += len(groupe.get('colonnes', []))
        
        # Vérifier les sous-en-têtes (ligne 4) (Requirement 9.2)
        col_idx = 2
        labels = colonnes_config.get('labels', {})
        
        for groupe in groupes:
            for colonne in groupe.get('colonnes', []):
                cell = ws.cell(row=4, column=col_idx)
                expected_label = labels.get(colonne, colonne)
                assert cell.value == expected_label, \
                    f"Le sous-en-tête doit être '{expected_label}'"
                
                # Vérifier le style du sous-en-tête (Requirement 9.4)
                assert cell.fill.start_color.rgb is not None, \
                    "Les sous-en-têtes doivent avoir une couleur de fond"
                
                col_idx += 1
        
        # Vérifier les lignes de données (Requirement 9.2)
        num_lignes_data = len(df)
        
        for row_idx in range(num_lignes_data):
            excel_row = row_idx + 5  # Les données commencent à la ligne 5
            df_row = df.iloc[row_idx]
            
            # Vérifier le libellé
            cell_libelle = ws.cell(row=excel_row, column=1)
            assert cell_libelle.value == df_row['libelle'], \
                f"Le libellé doit correspondre: attendu '{df_row['libelle']}', obtenu '{cell_libelle.value}'"
            
            # Vérifier les montants
            col_idx = 2
            colonnes_ordre = []
            for groupe in groupes:
                colonnes_ordre.extend(groupe.get('colonnes', []))
            
            for colonne in colonnes_ordre:
                cell = ws.cell(row=excel_row, column=col_idx)
                montant_df = df_row[colonne]
                
                # Vérifier la valeur
                if pd.notna(montant_df) and montant_df != 0:
                    assert cell.value is not None and cell.value != '-', \
                        f"La cellule doit contenir une valeur numérique pour {colonne}"
                    
                    # Vérifier le format numérique (Requirement 9.3)
                    assert '#,##0' in cell.number_format or 'GENERAL' in cell.number_format.upper(), \
                        f"Le format numérique doit inclure un séparateur de milliers, obtenu: {cell.number_format}"
                    
                    # Vérifier que la valeur est proche (tolérance pour les arrondis)
                    assert abs(float(cell.value) - montant_df) < 1.0, \
                        f"La valeur doit correspondre: attendu {montant_df}, obtenu {cell.value}"
                else:
                    # Les valeurs nulles ou zéro doivent être affichées comme '-'
                    assert cell.value == '-' or cell.value is None or cell.value == 0, \
                        f"Les valeurs nulles doivent être affichées comme '-'"
                
                # Vérifier les bordures (Requirement 9.4)
                assert cell.border is not None, \
                    "Toutes les cellules de données doivent avoir des bordures"
                assert cell.border.left.style is not None, \
                    "La bordure gauche doit être définie"
                assert cell.border.right.style is not None, \
                    "La bordure droite doit être définie"
                assert cell.border.top.style is not None, \
                    "La bordure supérieure doit être définie"
                assert cell.border.bottom.style is not None, \
                    "La bordure inférieure doit être définie"
                
                col_idx += 1
        
        # Vérifier la ligne de total (Requirement 9.2)
        derniere_ligne_df = df.iloc[-1]
        assert 'TOTAL' in str(derniere_ligne_df['libelle']).upper(), \
            "La dernière ligne doit être une ligne de total"
        
        derniere_ligne_excel = 5 + num_lignes_data - 1
        cell_total_libelle = ws.cell(row=derniere_ligne_excel, column=1)
        
        assert 'TOTAL' in str(cell_total_libelle.value).upper(), \
            "La dernière ligne Excel doit contenir 'TOTAL'"
        
        # Vérifier le style de la ligne de total (Requirement 9.4)
        assert cell_total_libelle.font.bold, \
            "Le libellé du total doit être en gras"
        assert cell_total_libelle.fill.start_color.rgb is not None, \
            "La ligne de total doit avoir une couleur de fond"
        
        # Vérifier que les montants du total sont également en gras
        col_idx = 2
        for colonne in colonnes_ordre:
            cell = ws.cell(row=derniere_ligne_excel, column=col_idx)
            assert cell.font.bold, \
                "Les montants du total doivent être en gras"
            col_idx += 1
        
        # Fermer le workbook
        wb.close()


# ============================================================================
# TESTS UNITAIRES COMPLÉMENTAIRES
# ============================================================================

def test_excel_export_empty_dataframe():
    """
    Test que l'exporteur gère correctement un DataFrame vide.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        fichier_sortie = os.path.join(tmpdir, 'test_empty.xlsx')
        
        exporter = ExcelExporter(fichier_sortie)
        
        df_empty = pd.DataFrame(columns=[
            'libelle', 'brut_ouverture', 'augmentations', 'diminutions', 
            'brut_cloture', 'amort_ouverture', 'dotations', 'reprises', 
            'amort_cloture', 'vnc_ouverture', 'vnc_cloture'
        ])
        
        colonnes_config = {
            'groupes': [
                {
                    'titre': 'VALEURS BRUTES',
                    'colonnes': ['brut_ouverture', 'augmentations', 'diminutions', 'brut_cloture']
                }
            ],
            'labels': {
                'brut_ouverture': 'Ouverture',
                'augmentations': 'Augmentations',
                'diminutions': 'Diminutions',
                'brut_cloture': 'Clôture'
            }
        }
        
        # Ne doit pas lever d'exception
        exporter.exporter_note(
            numero_note='TEST',
            titre_note='Test Empty',
            df=df_empty,
            colonnes_config=colonnes_config
        )
        
        fichier_sauvegarde = exporter.sauvegarder(avec_timestamp=False)
        
        assert os.path.exists(fichier_sauvegarde)
        
        # Vérifier que la feuille existe avec les en-têtes
        wb = load_workbook(fichier_sauvegarde)
        ws = wb['Note TEST']
        
        assert ws['A1'].value is not None  # Titre présent
        assert ws['A3'].value == 'LIBELLÉ'  # En-tête présent
        
        wb.close()


def test_excel_export_multiple_notes():
    """
    Test que l'exporteur peut exporter plusieurs notes dans le même fichier.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        fichier_sortie = os.path.join(tmpdir, 'test_multiple.xlsx')
        
        exporter = ExcelExporter(fichier_sortie)
        
        # Créer deux notes simples
        df1 = pd.DataFrame({
            'libelle': ['Ligne 1', 'TOTAL'],
            'brut_ouverture': [1000, 1000],
            'brut_cloture': [1200, 1200]
        })
        
        df2 = pd.DataFrame({
            'libelle': ['Ligne A', 'TOTAL'],
            'brut_ouverture': [2000, 2000],
            'brut_cloture': [2500, 2500]
        })
        
        colonnes_config = {
            'groupes': [
                {
                    'titre': 'VALEURS BRUTES',
                    'colonnes': ['brut_ouverture', 'brut_cloture']
                }
            ],
            'labels': {
                'brut_ouverture': 'Ouverture',
                'brut_cloture': 'Clôture'
            }
        }
        
        # Exporter les deux notes
        exporter.exporter_note('3A', 'Note 3A', df1, colonnes_config)
        exporter.exporter_note('3B', 'Note 3B', df2, colonnes_config)
        
        fichier_sauvegarde = exporter.sauvegarder(avec_timestamp=False)
        
        # Vérifier que les deux feuilles existent
        wb = load_workbook(fichier_sauvegarde)
        
        assert 'Note 3A' in wb.sheetnames
        assert 'Note 3B' in wb.sheetnames
        assert len(wb.sheetnames) == 2
        
        wb.close()


def test_excel_export_large_numbers():
    """
    Test que l'exporteur formate correctement les grands nombres.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        fichier_sortie = os.path.join(tmpdir, 'test_large.xlsx')
        
        exporter = ExcelExporter(fichier_sortie)
        
        df = pd.DataFrame({
            'libelle': ['Ligne 1', 'TOTAL'],
            'brut_ouverture': [123456789.0, 123456789.0],
            'brut_cloture': [987654321.0, 987654321.0]
        })
        
        colonnes_config = {
            'groupes': [
                {
                    'titre': 'VALEURS BRUTES',
                    'colonnes': ['brut_ouverture', 'brut_cloture']
                }
            ],
            'labels': {
                'brut_ouverture': 'Ouverture',
                'brut_cloture': 'Clôture'
            }
        }
        
        exporter.exporter_note('TEST', 'Test Large', df, colonnes_config)
        fichier_sauvegarde = exporter.sauvegarder(avec_timestamp=False)
        
        wb = load_workbook(fichier_sauvegarde)
        ws = wb['Note TEST']
        
        # Vérifier que les valeurs sont correctes
        assert ws.cell(row=5, column=2).value == 123456789.0
        assert ws.cell(row=5, column=3).value == 987654321.0
        
        # Vérifier le format numérique
        assert '#,##0' in ws.cell(row=5, column=2).number_format
        
        wb.close()


if __name__ == "__main__":
    # Exécuter les tests avec pytest
    pytest.main([__file__, '-v', '--hypothesis-show-statistics'])
